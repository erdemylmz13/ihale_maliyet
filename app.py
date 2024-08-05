from flask import Flask, request, render_template, send_file, jsonify
import os
import pandas as pd
from docx import Document
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from Levenshtein import distance as levenshtein_distance
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['BIRIM_FIYATLAR_PATH'] = 'C:/Users/ERDEM YILMAZ/Documents/my_project/2024-Birim-FIyatlar-Kitabi.csv'
app.config['STATIC_FOLDER'] = 'static'

# Birim fiyatları belleğe yükleyelim
birim_fiyatlar = None
vectorizer = None
tanim_vectors = None

def load_birim_fiyatlar():
    global birim_fiyatlar, vectorizer, tanim_vectors
    birim_fiyatlar = pd.read_csv(app.config['BIRIM_FIYATLAR_PATH'])
    birim_fiyatlar = birim_fiyatlar.rename(columns={
        'Poz No': 'Poz No',
        'Tanımı': 'Tanımı',
        'Ölçü Birimi': 'Ölçü Birimi',
        'Birim Fiyat': 'Birim Fiyat'
    })
    # NaN değerlerini boş string ile değiştir
    birim_fiyatlar['Tanımı'] = birim_fiyatlar['Tanımı'].fillna("")
    vectorizer, tanim_vectors = vectorize_descriptions(birim_fiyatlar)

def normalize_text(text):
    if not isinstance(text, str):
        text = str(text)
    text = text.lower()
    text = re.sub(r'-', ' ', text)  # Tire yerine boşluk koy
    text = re.sub(r'\s+', ' ', text)  # Birden fazla boşluğu tek boşluğa indir
    return re.sub(r'[^\w\s]', '', text)

def vectorize_descriptions(birim_fiyatlar):
    vectorizer = TfidfVectorizer()
    tanim_vectors = vectorizer.fit_transform(birim_fiyatlar['Tanımı'])
    return vectorizer, tanim_vectors

def levenshtein_similarity(tanim1, tanim2):
    distance = levenshtein_distance(tanim1, tanim2)
    max_len = max(len(tanim1), len(tanim2))
    return 1 - (distance / max_len)

def find_unit_price(tanim, poz_no, birim_fiyatlar, vectorizer, tanim_vectors, threshold=0.64, weight_lev=0.7):
    tanim_normalized = normalize_text(tanim)
    query_vector = vectorizer.transform([tanim_normalized])
    similarities = cosine_similarity(query_vector, tanim_vectors).flatten()
    
    best_matches = []
    # Poz No'ya göre tam eşleşme kontrolü yap
    if len(poz_no) >= 4:
        poz_matches = birim_fiyatlar[birim_fiyatlar['Poz No'].astype(str) == poz_no]
        if not poz_matches.empty:
            best_price = poz_matches.iloc[0]['Birim Fiyat']
            best_match = poz_matches.iloc[0]['Tanımı']
            best_poz_no = poz_matches.iloc[0]['Poz No']
            best_matches = [(1.0, best_price, best_match, best_poz_no, poz_matches.iloc[0]['Ölçü Birimi'])]
            eslesme_tanimi = f'<span style="color:green; font-weight:bold;">{best_match}</span>'
            return best_price, eslesme_tanimi, best_matches, best_poz_no

    for i, similarity in enumerate(similarities):
        if similarity > threshold:
            levenshtein_sim = levenshtein_similarity(tanim_normalized, normalize_text(birim_fiyatlar.iloc[i]['Tanımı']))
            combined_score = (similarity * (1 - weight_lev)) + (levenshtein_sim * weight_lev)
            best_matches.append((combined_score, birim_fiyatlar.iloc[i]['Birim Fiyat'], birim_fiyatlar.iloc[i]['Tanımı'], birim_fiyatlar.iloc[i]['Poz No'], birim_fiyatlar.iloc[i]['Ölçü Birimi']))

    best_matches.sort(reverse=True, key=lambda x: x[0])
    
    if best_matches:
        best_price_str = str(best_matches[0][1]).replace('.', '').replace(',', '.')
        best_price = float(best_price_str)
        best_match = best_matches[0][2]
        best_poz_no = best_matches[0][3]
        if best_matches[0][0] == 1.0:
            eslesme_tanimi = f'<span style="color:green; font-weight:bold;">{best_match}</span>'
        else:
            eslesme_tanimi = best_match
    else:
        best_price = 0.0
        eslesme_tanimi = "<span style='color: red; font-weight: bold;'>Eşleşme Bulunamadı. Birim Fiyata çift tıklayarak manuel giriş yapınız.</span>"
        best_poz_no = ""

    return best_price, eslesme_tanimi, best_matches[:3], best_poz_no

def extract_table_from_word(doc_path):
    doc = Document(doc_path)
    table = doc.tables[0]
    data = []
    for row in table.rows[2:]:
        cells = row.cells
        row_data = [
            cells[0].text.strip(),  # Sıra No
            cells[1].text.strip(),  # İş Kalemi No
            cells[2].text.strip(),  # İş Kaleminin Adı ve Kısa Açıklaması
            cells[3].text.strip(),  # Ölçü Birimi
            cells[4].text.strip().replace('.', '').replace(',', '.')   # Miktarı
        ]
        data.append(row_data)
    df = pd.DataFrame(data, columns=["Sıra No", "İş Kalemi No", "İş Kaleminin Adı ve Kısa Açıklaması", "Ölçü Birimi", "Miktarı"])
    df["Miktarı"] = pd.to_numeric(df["Miktarı"], errors='coerce').fillna(0)
    return df

def update_cost_table(df, birim_fiyatlar, vectorizer, tanim_vectors):
    df["Birim Fiyat (TL)"] = 0
    df["Tutar (TL)"] = 0
    df["Bulunulan Poz No"] = ""
    df["Eşleşme Durumu"] = ""
    df["Alternatifler"] = [[] for _ in range(len(df))]

    for idx, row in df.iterrows():
        tanim = row["İş Kaleminin Adı ve Kısa Açıklaması"]
        poz_no = row["İş Kalemi No"]

        birim_fiyat, eslesme_tanimi, best_matches, best_poz_no = find_unit_price(tanim, poz_no, birim_fiyatlar, vectorizer, tanim_vectors)

        if any("color:green" in alt[2] for alt in best_matches):
            eslesme_tanimi = f'<span style="color:green;">{eslesme_tanimi}</span>'
        df.at[idx, "Birim Fiyat (TL)"] = birim_fiyat
        df.at[idx, "Tutar (TL)"] = row["Miktarı"] * birim_fiyat
        df.at[idx, "Bulunulan Poz No"] = best_poz_no if best_poz_no else ""
        df.at[idx, "Eşleşme Durumu"] = eslesme_tanimi
        df.at[idx, "Alternatifler"] = best_matches

    df["Miktarı"] = df["Miktarı"].apply(lambda x: f"{x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
    df["Birim Fiyat (TL)"] = df["Birim Fiyat (TL)"].apply(lambda x: f"{x:,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.') if pd.notnull(x) else "")
    df["Tutar (TL)"] = df["Tutar (TL)"].apply(lambda x: f"{x:,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.') if pd.notnull(x) else "")

    total_cost = df["Tutar (TL)"].str.replace(' TL', '').str.replace('.', '').str.replace(',', '.').astype(float).sum()

    total_row = pd.DataFrame([["", "", "Toplam", "", "", "", f"{total_cost:,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.'), "", "", ""]], columns=df.columns)
    df = pd.concat([df, total_row], ignore_index=True)

    df = df[~((df["İş Kaleminin Adı ve Kısa Açıklaması"] == "TOPLAM TUTAR(K.D.V Hariç)") & (df["Miktarı"] == "0,00") & (df["Tutar (TL)"] == "0,00 TL"))]
    df = df[~((df["İş Kalemi No"] == "") & (df["İş Kaleminin Adı ve Kısa Açıklaması"] == "Toplam"))]

    return df, f"{total_cost:,.2f} TL".replace(',', 'X').replace('.', ',').replace('X', '.')

def save_to_excel(df, path):
    # Alternatifler sütununu kaldır
    if "Alternatifler" in df.columns:
        df = df.drop(columns=["Alternatifler"])

    # Excel çalışma kitabı ve çalışma sayfası oluşturma
    wb = Workbook()
    ws = wb.active

    # DataFrame'i satırlara dönüştür ve çalışma sayfasına ekle
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)
        if r_idx == 0:  # Başlık satırı için stil ekleme
            for c_idx, cell in enumerate(ws[r_idx + 1], 1):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=Side(border_style="thin", color="000000"),
                                     right=Side(border_style="thin", color="000000"),
                                     top=Side(border_style="thin", color="000000"),
                                     bottom=Side(border_style="thin", color="000000"))
                cell.font = Font(bold=True, size=12)  # Başlıkları kalın ve 1 punto büyük yap
        else:  # Diğer satırlar için stil ekleme
            for c_idx, cell in enumerate(ws[r_idx + 1], 1):
                cell.border = Border(left=Side(border_style="thin", color="000000"),
                                     right=Side(border_style="thin", color="000000"),
                                     top=Side(border_style="thin", color="000000"),
                                     bottom=Side(border_style="thin", color="000000"))
                if c_idx == 1:  # Sıra numaralarını kalın yap
                    cell.font = Font(bold=True)
                if c_idx == 3 or c_idx == 8:  # Tanımı ve Eşleşme Durumu sütunları için metin kaydırma
                    cell.alignment = Alignment(wrap_text=True)
                if isinstance(cell.value, str) and 'color: red;' in cell.value:
                    cell.font = Font(color="FF0000", bold=True)
                    cell.value = re.sub(r'<.*?>', '', cell.value)
                elif isinstance(cell.value, str) and 'color:green;' in cell.value:
                    cell.font = Font(color="00FF00", bold=True)
                    cell.value = re.sub(r'<.*?>', '', cell.value)

    # Sütun genişliklerini ayarla
    ws.column_dimensions['C'].width = 50  # İş Kaleminin Adı ve Kısa Açıklaması sütunu
    ws.column_dimensions['I'].width = 50  # Eşleşme Durumu sütunu

    # Otomatik sütun genişliği ayarı
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) if column not in ['C', 'I'] else 50
        ws.column_dimensions[column].width = adjusted_width

    # Çalışma kitabını kaydet
    wb.save(path)

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(file_path)
            table_df = extract_table_from_word(file_path)
            load_birim_fiyatlar()  # Birim fiyatları yükleyelim
            updated_df, total_cost = update_cost_table(table_df, birim_fiyatlar, vectorizer, tanim_vectors)
            excel_output_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output.xlsx')
            save_to_excel(updated_df, excel_output_path)
            ihale_kayit_numarasi = '_'.join(file.filename.split('_')[0:2])
            rows = updated_df.to_dict(orient='records')
            return render_template('result.html', rows=rows, total_cost=total_cost, excel_path=excel_output_path, ihale_kayit_numarasi=ihale_kayit_numarasi)
    return render_template('upload.html')

@app.route('/download_excel')
def download_excel():
    excel_path = request.args.get('path')
    return send_file(excel_path, as_attachment=True)

@app.route('/update_unit_price', methods=['POST'])
def update_unit_price():
    data = request.get_json()
    row_id = int(data['rowId'])
    new_unit_price = float(data['unitPrice'].replace(',', '.'))

    return jsonify(status="success")

@app.route('/search', methods=['GET', 'POST'])
def search():
    global birim_fiyatlar, vectorizer, tanim_vectors
    if birim_fiyatlar is None or vectorizer is None or tanim_vectors is None:
        load_birim_fiyatlar()
        
    if request.method == 'POST':
        query = request.form['query']
        query_normalized = normalize_text(query)
        query_vector = vectorizer.transform([query_normalized])
        similarities = cosine_similarity(query_vector, tanim_vectors).flatten()
        
        # Poz numarası ile eşleşmeleri bul
        poz_matches = birim_fiyatlar[birim_fiyatlar['Poz No'].astype(str).str.contains(query, case=False, na=False)]

        # Tanım ile eşleşmeleri bul
        best_matches = []
        for i, similarity in enumerate(similarities):
            if similarity > 0:
                levenshtein_sim = levenshtein_similarity(query_normalized, normalize_text(birim_fiyatlar.iloc[i]['Tanımı']))
                combined_score = (similarity + levenshtein_sim) / 2
                best_matches.append((combined_score, birim_fiyatlar.iloc[i]['Poz No'], birim_fiyatlar.iloc[i]['Tanımı'], birim_fiyatlar.iloc[i]['Ölçü Birimi'], birim_fiyatlar.iloc[i]['Birim Fiyat']))
        
        best_matches.sort(reverse=True, key=lambda x: x[0])
        results = pd.DataFrame(best_matches, columns=['Eşleşme Oranı', 'Poz No', 'Tanımı', 'Ölçü Birimi', 'Birim Fiyat'])
        
        results['Eşleşme Oranı'] = (results['Eşleşme Oranı'] * 100).round(2).astype(str) + '%'
        
        def clean_and_format_price(x):
            if isinstance(x, str):
                x = x.replace('\n', '').replace('.', '').replace(',', '.')
            return f"{float(x):,.2f} TL".replace('.', 'X').replace(',', '.').replace('X', ',')

        results['Birim Fiyat'] = results['Birim Fiyat'].apply(lambda x: clean_and_format_price(x) if pd.notnull(x) else "")
        
        if not poz_matches.empty:
            poz_matches['Eşleşme Oranı'] = '100%'
            poz_matches['Birim Fiyat'] = poz_matches['Birim Fiyat'].apply(lambda x: clean_and_format_price(x) if pd.notnull(x) else "")
            results = pd.concat([poz_matches[['Eşleşme Oranı', 'Poz No', 'Tanımı', 'Ölçü Birimi', 'Birim Fiyat']], results], ignore_index=True)
        
        return render_template('search_results.html', query=query, results=results.to_html(index=False, escape=False))
    return render_template('upload.html')

if __name__ == '__main__':
    app.run(debug=True)

